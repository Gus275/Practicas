import openpyxl
from openpyxl import Workbook

workbook = Workbook()
  

workbook.save(filename="demo.xlsx")

wb = openpyxl.Workbook()
productos = wb.active
a1 = productos.cell(row = 1, column = 1)
a2 = productos.cell(row = 1, column = 2)
a3 = productos.cell(row = 1, column = 3)
b1 = productos.cell(row = 2, column = 1)
b2 = productos.cell(row = 2, column = 2)
b3 = productos.cell(row = 2, column = 3)
c1 = productos.cell(row = 3, column = 1)
c2 = productos.cell(row = 3, column = 2)
c3 = productos.cell(row = 3, column = 3)

a1.value = "SKU"
b1.value = "Nombre"
c1.value = "Unidad"

wb.save("demo.xlsx") 

# obtener los valores de las celdas
nombre =productos['a1'].value
edad = productos['b1'].value

# imprimir los valores obtenidos
print(f'Nombre: {nombre}')
print(f'Edad: {edad}')